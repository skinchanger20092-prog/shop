import asyncio
import json
import logging
import os
import re
import secrets
import shutil
import sqlite3
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin

import aiohttp
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import FSInputFile, KeyboardButton, Message, ReplyKeyboardMarkup
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage


load_dotenv()

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = 7541282558
BASE_DIR = Path(__file__).resolve().parent
DB_FILE = BASE_DIR / "bot.db"
USERS_FILE = BASE_DIR / "users.json"
WELCOME_IMAGE = BASE_DIR / "Dobro.png"
ORDER_IMAGE = BASE_DIR / "strana.png"
TARIFF_IMAGE = BASE_DIR / "tarif.png"
SUPPORT_IMAGE = BASE_DIR / "teh.png"
INFO_IMAGE = BASE_DIR / "info.png"
YUAN_IMAGE = BASE_DIR / "kurs.png"
CARGO_ORDERS_DIR = BASE_DIR / "cargo_orders"
CARGO_PHOTOS_DIR = BASE_DIR / "cargo_photos"
CARGO_TEMPLATE = BASE_DIR / "каргос.xlsx"
SEARCH_SOURCES = [
    "https://www.goofish.com/",
    "https://mxjstore.x.yupoo.com/albums",
    "https://wwfake100.x.yupoo.com/albums",
    "https://yolo66.x.yupoo.com/albums",
    "https://vx798134596.x.yupoo.com/",
    "https://konng-gonng.x.yupoo.com/",
    "https://goat-official.x.yupoo.com/categories",
    "https://huskyreps.x.yupoo.com/",
    "https://powerball.x.yupoo.com/",
    "https://anniestudio.x.yupoo.com/",
    "https://repsking.x.yupoo.com/",
    "https://loganhere.x.yupoo.com/",
    "https://martinreps.x.yupoo.com/",
    "https://windvane168.x.yupoo.com/",
    "https://lufilostudio.x.yupoo.com/",
    "https://rainbowreps.x.yupoo.com/",
    "https://angelking47.x.yupoo.com/",
    "https://pikachushop.x.yupoo.com/",
]
HTTP_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "ru,en-US;q=0.9,en;q=0.8",
}
REQUEST_TIMEOUT = aiohttp.ClientTimeout(total=12)
MAX_RESULTS = 8
MAX_MESSAGES_PER_WINDOW = 8
RATE_WINDOW_SECONDS = 10
RATE_LIMIT_NOTICE_COOLDOWN = 5
TRACKING_STAGES = {
    "1": "Выкуплен",
    "2": "Приехал на склад",
    "3": "Отправлен",
    "4": "Приехал в РФ",
}
ACTIVE_ORDER_STATUSES = {
    "Оформляет заказ",
    "Отправляет данные",
    "Заявка отправлена",
    "Заказ обработан",
}

if not TOKEN:
    raise RuntimeError("Не найден BOT_TOKEN. Добавь токен в файл .env.")


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)

dp = Dispatcher()
china_order_users: set[int] = set()
admin_search_users: set[int] = set()
paid_search_users: set[int] = set()
tracking_lookup_users: set[int] = set()
search_semaphore = asyncio.Semaphore(3)
user_request_times: dict[int, list[float]] = {}
user_last_limit_notice: dict[int, float] = {}
admin_excel_sessions: dict[int, dict] = {}

user_main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Сделать заказ")],
        [KeyboardButton(text="Тарифы")],
        [KeyboardButton(text="Курс юаня")],
        [KeyboardButton(text="Инфо")],
        [KeyboardButton(text="Тех. поддержка")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Напиши сообщение или выбери кнопку",
)

admin_main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Сделать заказ")],
        [KeyboardButton(text="Найти эту вещь в Китае")],
        [KeyboardButton(text="Неодобренные заказы")],
        [KeyboardButton(text="Все заказы и их статусы")],
        [KeyboardButton(text="Тарифы")],
        [KeyboardButton(text="Курс юаня")],
        [KeyboardButton(text="Инфо")],
        [KeyboardButton(text="Тех. поддержка")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Напиши сообщение или выбери кнопку",
)

user_order_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Китай")],
        [KeyboardButton(text="Найти эту вещь в Китае")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выбери раздел",
)

china_submit_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Я все скинул")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Отправь данные по заказу или нажми кнопку ниже",
)

payment_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Оплатить заказ")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Нажми кнопку для оплаты заказа",
)

china_tariff_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Китай-Благовещенск-Керчь (2-5 дней до РФ, Владивостока)")],
        [KeyboardButton(text="Китай-Москва-Керчь (20-30 дней до РФ, МСК)")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выбери тариф доставки",
)

info_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Мой заказ")],
        [KeyboardButton(text="Мои попытки поиска")],
        [KeyboardButton(text="Узнать статус заказа")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выбери действие",
)


user_main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Сделать заказ")],
        [KeyboardButton(text="Тарифы")],
        [KeyboardButton(text="Курс юаня")],
        [KeyboardButton(text="Инфо")],
        [KeyboardButton(text="Тех. поддержка")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Напиши сообщение или выбери кнопку",
)

admin_main_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Сделать заказ")],
        [KeyboardButton(text="Найти эту вещь в Китае")],
        [KeyboardButton(text="Неодобренные заказы")],
        [KeyboardButton(text="Все заказы и их статусы")],
        [KeyboardButton(text="Тарифы")],
        [KeyboardButton(text="Курс юаня")],
        [KeyboardButton(text="Инфо")],
        [KeyboardButton(text="Тех. поддержка")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Напиши сообщение или выбери кнопку",
)

user_order_keyboard = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="Китай")],
        [KeyboardButton(text="Найти эту вещь в Китае")],
        [KeyboardButton(text="Назад")],
    ],
    resize_keyboard=True,
    input_field_placeholder="Выбери раздел",
)


def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def get_db() -> sqlite3.Connection:
    connection = sqlite3.connect(DB_FILE)
    connection.row_factory = sqlite3.Row
    return connection


def init_db() -> None:
    with get_db() as connection:
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
            """
        )
        connection.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY,
                first_name TEXT NOT NULL DEFAULT '',
                last_name TEXT NOT NULL DEFAULT '',
                username TEXT NOT NULL DEFAULT '',
                first_seen TEXT NOT NULL DEFAULT '',
                last_seen TEXT NOT NULL DEFAULT '',
                country TEXT NOT NULL DEFAULT '',
                tariff TEXT NOT NULL DEFAULT '',
                status TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT '',
                discount INTEGER NOT NULL DEFAULT 0,
                buyout_free INTEGER NOT NULL DEFAULT 0,
                is_banned INTEGER NOT NULL DEFAULT 0,
                ban_reason TEXT NOT NULL DEFAULT '',
                payment_status TEXT NOT NULL DEFAULT '',
                search_access INTEGER NOT NULL DEFAULT 0,
                order_number INTEGER NOT NULL DEFAULT 0,
                tracking_code TEXT NOT NULL DEFAULT '',
                tracking_stage TEXT NOT NULL DEFAULT ''
            )
            """
        )
        connection.execute(
            """
            CREATE UNIQUE INDEX IF NOT EXISTS idx_users_tracking_code
            ON users(tracking_code)
            WHERE tracking_code <> ''
            """
        )
        columns = {
            row["name"]
            for row in connection.execute("PRAGMA table_info(users)").fetchall()
        }
        if "search_access" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN search_access INTEGER NOT NULL DEFAULT 0"
            )
        if "tariff" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN tariff TEXT NOT NULL DEFAULT ''"
            )
        if "discount" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN discount INTEGER NOT NULL DEFAULT 0"
            )
        if "buyout_free" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN buyout_free INTEGER NOT NULL DEFAULT 0"
            )
        if "is_banned" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN is_banned INTEGER NOT NULL DEFAULT 0"
            )
        if "ban_reason" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN ban_reason TEXT NOT NULL DEFAULT ''"
            )
        if "payment_status" not in columns:
            connection.execute(
                "ALTER TABLE users ADD COLUMN payment_status TEXT NOT NULL DEFAULT ''"
            )


def get_setting(key: str, default: str = "") -> str:
    with get_db() as connection:
        row = connection.execute(
            "SELECT value FROM settings WHERE key = ?",
            (key,),
        ).fetchone()
    if not row:
        return default
    return row["value"] or default


def set_setting(key: str, value: str) -> None:
    with get_db() as connection:
        connection.execute(
            """
            INSERT INTO settings (key, value)
            VALUES (?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (key, value),
        )


def get_yuan_rate() -> str:
    return get_setting("yuan_rate", "13,0")


def set_yuan_rate(value: str) -> None:
    set_setting("yuan_rate", value)


def ensure_cargo_dirs() -> None:
    CARGO_ORDERS_DIR.mkdir(parents=True, exist_ok=True)
    CARGO_PHOTOS_DIR.mkdir(parents=True, exist_ok=True)


def get_active_cargo_file() -> Path | None:
    raw = get_setting("active_cargo_file", "").strip()
    if not raw:
        return None
    path = Path(raw)
    return path if path.exists() else None


def set_active_cargo_file(path: Path | None) -> None:
    set_setting("active_cargo_file", str(path) if path else "")


def list_cargo_files() -> list[Path]:
    ensure_cargo_dirs()
    return sorted(CARGO_ORDERS_DIR.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)


def create_cargo_file() -> Path:
    ensure_cargo_dirs()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cargo_{timestamp}.xlsx"
    target = CARGO_ORDERS_DIR / filename
    shutil.copy2(CARGO_TEMPLATE, target)
    set_active_cargo_file(target)
    return target


def resolve_cargo_file(value: str | None) -> Path | None:
    ensure_cargo_dirs()
    target = (value or "current").strip()
    if not target or target.lower() == "current":
        return get_active_cargo_file()
    files = list_cargo_files()
    if target.isdigit():
        index = int(target) - 1
        if 0 <= index < len(files):
            return files[index]
        return None
    path = CARGO_ORDERS_DIR / target
    return path if path.exists() else None


def get_cargo_next_row(sheet) -> int:
    row = 5
    while sheet[f"C{row}"].value or sheet[f"D{row}"].value or sheet[f"E{row}"].value:
        row += 1
    return row


def append_product_to_cargo(file_path: Path, product: dict) -> int:
    workbook = load_workbook(file_path)
    sheet = workbook.active
    row = get_cargo_next_row(sheet)
    sheet[f"C{row}"] = product.get("name", "")
    sheet[f"D{row}"] = product.get("link", "")
    sheet[f"E{row}"] = f"Размер: {product.get('size', '')}\nЦвет: {product.get('color', '')}"
    sheet[f"F{row}"] = product.get("price", "")
    sheet[f"G{row}"] = product.get("quantity", "")
    try:
        price = float(str(product.get("price", "0")).replace(",", "."))
        quantity = float(str(product.get("quantity", "0")).replace(",", "."))
        sheet[f"H{row}"] = price * quantity
    except ValueError:
        sheet[f"H{row}"] = ""
    sheet[f"I{row}"] = product.get("delivery", "")
    if product.get("photo_path"):
        try:
            image = XLImage(str(product["photo_path"]))
            image.width = 90
            image.height = 90
            sheet.add_image(image, f"J{row}")
        except Exception:
            pass
    workbook.save(file_path)
    return row


def read_cargo_rows(file_path: Path) -> list[dict]:
    workbook = load_workbook(file_path)
    sheet = workbook.active
    rows: list[dict] = []
    row = 5
    while row <= sheet.max_row:
        name = sheet[f"C{row}"].value
        link = sheet[f"D{row}"].value
        specs = sheet[f"E{row}"].value
        price = sheet[f"F{row}"].value
        quantity = sheet[f"G{row}"].value
        total = sheet[f"H{row}"].value
        delivery = sheet[f"I{row}"].value
        if any(value not in (None, "") for value in [name, link, specs, price, quantity, total, delivery]):
            rows.append(
                {
                    "row": row,
                    "name": name or "",
                    "link": link or "",
                    "specs": specs or "",
                    "price": price or "",
                    "quantity": quantity or "",
                    "total": total or "",
                    "delivery": delivery or "",
                }
            )
        row += 1
    return rows


def is_maintenance_mode() -> bool:
    return get_setting("maintenance_mode", "0") == "1"


def set_maintenance_mode(enabled: bool) -> None:
    set_setting("maintenance_mode", "1" if enabled else "0")


def row_to_user(row: sqlite3.Row) -> dict:
    return {
        "id": row["id"],
        "first_name": row["first_name"] or "",
        "last_name": row["last_name"] or "",
        "username": row["username"] or "",
        "first_seen": row["first_seen"] or "",
        "last_seen": row["last_seen"] or "",
        "country": row["country"] or "",
        "tariff": row["tariff"] if "tariff" in row.keys() else "",
        "status": row["status"] or "",
        "note": row["note"] or "",
        "discount": int(row["discount"] or 0) if "discount" in row.keys() else 0,
        "buyout_free": int(row["buyout_free"] or 0) if "buyout_free" in row.keys() else 0,
        "is_banned": int(row["is_banned"] or 0) if "is_banned" in row.keys() else 0,
        "ban_reason": row["ban_reason"] if "ban_reason" in row.keys() else "",
        "payment_status": row["payment_status"] if "payment_status" in row.keys() else "",
        "search_access": int(row["search_access"] or 0),
        "order_number": int(row["order_number"] or 0),
        "tracking_code": row["tracking_code"] or "",
        "tracking_stage": row["tracking_stage"] or "",
    }


def load_legacy_users() -> dict[str, dict]:
    if not USERS_FILE.exists():
        return {}

    try:
        data = json.loads(USERS_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (json.JSONDecodeError, OSError):
        logging.warning("Не удалось прочитать users.json, пропускаю миграцию.")
        return {}


def migrate_legacy_users() -> None:
    legacy_users = load_legacy_users()
    if not legacy_users:
        return

    with get_db() as connection:
        existing_count = connection.execute("SELECT COUNT(*) FROM users").fetchone()[0]
        if existing_count:
            return

        for raw_user in legacy_users.values():
            user_id = int(raw_user.get("id", 0) or 0)
            if not user_id:
                continue

            connection.execute(
                """
                INSERT OR REPLACE INTO users (
                    id, first_name, last_name, username,
                    first_seen, last_seen, country, tariff, status,
                    note, discount, buyout_free, is_banned, ban_reason, payment_status,
                    search_access, order_number, tracking_code, tracking_stage
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    user_id,
                    raw_user.get("first_name", "") or "",
                    raw_user.get("last_name", "") or "",
                    raw_user.get("username", "") or "",
                    raw_user.get("first_seen", "") or now_str(),
                    raw_user.get("last_seen", "") or now_str(),
                    raw_user.get("country", "") or "",
                    raw_user.get("tariff", "") or "",
                    raw_user.get("status", "Новый") or "Новый",
                    raw_user.get("note", "") or "",
                    int(raw_user.get("discount", 0) or 0),
                    int(raw_user.get("buyout_free", 0) or 0),
                    int(raw_user.get("is_banned", 0) or 0),
                    raw_user.get("ban_reason", "") or "",
                    raw_user.get("payment_status", "") or "",
                    int(raw_user.get("search_access", 0) or 0),
                    int(raw_user.get("order_number", 0) or 0),
                    raw_user.get("tracking_code", "") or "",
                    raw_user.get("tracking_stage", "") or "",
                ),
            )


def load_users() -> dict[str, dict]:
    try:
        with get_db() as connection:
            rows = connection.execute("SELECT * FROM users ORDER BY id").fetchall()
    except sqlite3.OperationalError as error:
        if "no such table: users" not in str(error).lower():
            raise
        init_db()
        with get_db() as connection:
            rows = connection.execute("SELECT * FROM users ORDER BY id").fetchall()
    return {str(row["id"]): row_to_user(row) for row in rows}


def save_users(users: dict[str, dict]) -> None:
    with get_db() as connection:
        for user in users.values():
            user_id = int(user.get("id", 0) or 0)
            if not user_id:
                continue

            connection.execute(
                """
                INSERT INTO users (
                    id, first_name, last_name, username,
                    first_seen, last_seen, country, tariff, status,
                    note, discount, buyout_free, is_banned, ban_reason, payment_status,
                    search_access, order_number, tracking_code, tracking_stage
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(id) DO UPDATE SET
                    first_name=excluded.first_name,
                    last_name=excluded.last_name,
                    username=excluded.username,
                    first_seen=excluded.first_seen,
                    last_seen=excluded.last_seen,
                    country=excluded.country,
                    tariff=excluded.tariff,
                    status=excluded.status,
                    note=excluded.note,
                    discount=excluded.discount,
                    buyout_free=excluded.buyout_free,
                    is_banned=excluded.is_banned,
                    ban_reason=excluded.ban_reason,
                    payment_status=excluded.payment_status,
                    search_access=excluded.search_access,
                    order_number=excluded.order_number,
                    tracking_code=excluded.tracking_code,
                    tracking_stage=excluded.tracking_stage
                """,
                (
                    user_id,
                    user.get("first_name", "") or "",
                    user.get("last_name", "") or "",
                    user.get("username", "") or "",
                    user.get("first_seen", "") or now_str(),
                    user.get("last_seen", "") or now_str(),
                    user.get("country", "") or "",
                    user.get("tariff", "") or "",
                    user.get("status", "Новый") or "Новый",
                    user.get("note", "") or "",
                    int(user.get("discount", 0) or 0),
                    int(user.get("buyout_free", 0) or 0),
                    int(user.get("is_banned", 0) or 0),
                    user.get("ban_reason", "") or "",
                    user.get("payment_status", "") or "",
                    int(user.get("search_access", 0) or 0),
                    int(user.get("order_number", 0) or 0),
                    user.get("tracking_code", "") or "",
                    user.get("tracking_stage", "") or "",
                ),
            )

def update_user(
    message: Message,
    *,
    country: str | None = None,
    tariff: str | None = None,
    status: str | None = None,
) -> dict | None:
    if not message.from_user:
        return None

    users = load_users()
    user_id = str(message.from_user.id)
    existing = users.get(user_id, {})

    user_data = {
        "id": message.from_user.id,
        "first_name": message.from_user.first_name or "",
        "last_name": message.from_user.last_name or "",
        "username": message.from_user.username or "",
        "first_seen": existing.get("first_seen", now_str()),
        "last_seen": now_str(),
        "country": country if country is not None else existing.get("country", ""),
        "tariff": tariff if tariff is not None else existing.get("tariff", ""),
        "status": status if status is not None else existing.get("status", "Новый"),
        "note": existing.get("note", ""),
        "discount": int(existing.get("discount", 0) or 0),
        "buyout_free": int(existing.get("buyout_free", 0) or 0),
        "is_banned": int(existing.get("is_banned", 0) or 0),
        "ban_reason": existing.get("ban_reason", ""),
        "payment_status": existing.get("payment_status", ""),
        "search_access": int(existing.get("search_access", 0) or 0),
        "order_number": existing.get("order_number", 0),
        "tracking_code": existing.get("tracking_code", ""),
        "tracking_stage": existing.get("tracking_stage", ""),
    }

    users[user_id] = user_data
    save_users(users)
    return user_data


def get_user(user_id: int) -> dict | None:
    return load_users().get(str(user_id))


def update_user_fields(user_id: int, **fields) -> dict | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    for field, value in fields.items():
        users[key][field] = value
    users[key]["last_seen"] = now_str()
    save_users(users)
    return users[key]


def get_user_by_tracking_code(tracking_code: str) -> tuple[str | None, dict | None]:
    users = load_users()
    normalized = tracking_code.strip().upper()
    for key, user in users.items():
        if (user.get("tracking_code", "") or "").upper() == normalized:
            return key, user
    return None, None


def get_user_by_order_number(order_number: int) -> dict | None:
    users = load_users()
    for user in users.values():
        if int(user.get("order_number", 0) or 0) == order_number:
            return user
    return None


def set_user_status(user_id: int, status: str) -> None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return
    users[key]["status"] = status
    users[key]["last_seen"] = now_str()
    save_users(users)


def set_user_note(user_id: int, note: str) -> bool:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return False
    users[key]["note"] = note
    users[key]["last_seen"] = now_str()
    save_users(users)
    return True


def get_banned_users() -> list[dict]:
    return [user for user in load_users().values() if int(user.get("is_banned", 0) or 0) == 1]


def assign_order_number(user_id: int) -> int | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    last_number = max((int(user.get("order_number", 0)) for user in users.values()), default=0)
    new_number = last_number + 1
    users[key]["order_number"] = new_number
    users[key]["tracking_code"] = ""
    users[key]["tracking_stage"] = ""
    users[key]["payment_status"] = ""
    users[key]["last_seen"] = now_str()
    save_users(users)
    return new_number


def can_approve_order(user: dict | None) -> bool:
    if not user:
        return False
    return user.get("status") in {"Заявка отправлена", "Отправляет данные"}


def can_cancel_order(user: dict | None) -> bool:
    if not user:
        return False
    if user.get("status") == "Заказ отменен":
        return False
    return bool(
        user.get("order_number", 0)
        or user.get("tracking_code")
        or user.get("status") in {
            "Оформляет заказ",
            "Отправляет данные",
            "Заявка отправлена",
            "Заказ обработан",
        }
    )


def get_pending_orders() -> list[dict]:
    users = load_users()
    pending = [
        user for user in users.values()
        if user.get("status") in {"Заявка отправлена", "Отправляет данные"}
    ]
    pending.sort(key=lambda item: item.get("last_seen", ""), reverse=True)
    return pending


def get_all_orders() -> list[dict]:
    users = load_users()
    orders = [
        user for user in users.values()
        if (
            user.get("order_number", 0)
            or user.get("tracking_code")
            or user.get("status") in ACTIVE_ORDER_STATUSES
        )
    ]
    orders.sort(
        key=lambda item: (
            int(item.get("order_number", 0) or 0),
            item.get("last_seen", ""),
        ),
        reverse=True,
    )
    return orders


def get_active_orders() -> list[dict]:
    return [
        user for user in get_all_orders()
        if user.get("tracking_stage") != TRACKING_STAGES["4"]
    ]


def get_tracked_orders() -> list[dict]:
    orders = [user for user in get_all_orders() if user.get("tracking_code")]
    orders.sort(key=lambda item: item.get("last_seen", ""), reverse=True)
    return orders


def get_recent_users(limit: int = 10) -> list[dict]:
    users = load_users()
    recent = sorted(users.values(), key=lambda item: item.get("last_seen", ""), reverse=True)
    return recent[:limit]


def get_today_stats() -> dict[str, int]:
    today_prefix = datetime.now().strftime("%Y-%m-%d")
    users = list(load_users().values())
    return {
        "new_users": sum(1 for user in users if (user.get("first_seen", "") or "").startswith(today_prefix)),
        "active_today": sum(1 for user in users if (user.get("last_seen", "") or "").startswith(today_prefix)),
        "orders_today": sum(
            1
            for user in users
            if int(user.get("order_number", 0) or 0) > 0 and (user.get("last_seen", "") or "").startswith(today_prefix)
        ),
        "processed_today": sum(
            1
            for user in users
            if user.get("status") == "Заказ обработан" and (user.get("last_seen", "") or "").startswith(today_prefix)
        ),
    }


def get_search_access_users() -> list[dict]:
    users = load_users()
    access_users = [user for user in users.values() if int(user.get("search_access", 0) or 0) > 0]
    access_users.sort(key=lambda item: item.get("last_seen", ""), reverse=True)
    return access_users


def search_users(query: str, limit: int = 12) -> list[dict]:
    normalized_query = normalize_query(query)
    if not normalized_query:
        return []

    results: list[dict] = []
    for user in load_users().values():
        haystack = normalize_query(
            " ".join(
                [
                    str(user.get("id", "")),
                    user.get("first_name", "") or "",
                    user.get("last_name", "") or "",
                    user.get("username", "") or "",
                    user.get("country", "") or "",
                    user.get("status", "") or "",
                    user.get("note", "") or "",
                    str(user.get("order_number", 0) or ""),
                    user.get("tracking_code", "") or "",
                    user.get("tracking_stage", "") or "",
                ]
            )
        )
        if normalized_query in haystack:
            results.append(user)

    results.sort(key=lambda item: item.get("last_seen", ""), reverse=True)
    return results[:limit]


def add_search_credits(user_id: int, amount: int) -> int | None:
    users = load_users()
    key = str(user_id)
    if key not in users or amount <= 0:
        return None

    current = int(users[key].get("search_access", 0) or 0)
    users[key]["search_access"] = current + amount
    users[key]["last_seen"] = now_str()
    save_users(users)
    return int(users[key]["search_access"])


def set_search_credits(user_id: int, amount: int) -> int | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    users[key]["search_access"] = max(0, int(amount))
    users[key]["last_seen"] = now_str()
    save_users(users)
    return int(users[key]["search_access"])


def consume_search_credit(user_id: int) -> int | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    current = int(users[key].get("search_access", 0) or 0)
    if current <= 0:
        return None

    users[key]["search_access"] = current - 1
    users[key]["last_seen"] = now_str()
    save_users(users)
    return int(users[key]["search_access"])


def format_order_card(user: dict, *, include_action_hint: bool = False) -> str:
    username = f"@{user['username']}" if user.get("username") else "без username"
    full_name = " ".join(
        part for part in [user.get("first_name", ""), user.get("last_name", "")] if part
    ).strip() or "Без имени"

    lines = [
        f"Заказ №{user.get('order_number', 0) or '-'}",
        f"ID: {user['id']}",
        f"Имя: {full_name}",
        f"Username: {username}",
        f"Страна: {user.get('country', '') or 'Не выбрана'}",
        f"Тариф: {user.get('tariff', '') or 'Не выбран'}",
        f"Статус: {user.get('status', '') or 'Новый'}",
        f"Скидка: {int(user.get('discount', 0) or 0)}%",
        f"Выкуп без комиссии: {'Да' if int(user.get('buyout_free', 0) or 0) == 1 else 'Нет'}",
        f"Бан: {'Да' if int(user.get('is_banned', 0) or 0) == 1 else 'Нет'}",
        f"Попыток поиска: {int(user.get('search_access', 0) or 0)}",
        f"Трек-код: {user.get('tracking_code', '') or 'Нет'}",
        f"Этап: {user.get('tracking_stage', '') or 'Нет'}",
        f"Обновлен: {user.get('last_seen', '-')}",
    ]
    if user.get("ban_reason"):
        lines.append(f"Причина бана: {user['ban_reason']}")
    if user.get("payment_status"):
        lines.append(f"Оплата: {user['payment_status']}")

    if include_action_hint and user.get("status") in {"Заявка отправлена", "Отправляет данные"}:
        lines.append(f"Для одобрения: /done {user['id']}")
    if include_action_hint and user.get("status") != "Заказ отменен":
        lines.append(f"Для отмены: /cancel {user['id']}")

    return "\n".join(lines)


def generate_tracking_code() -> str:
    random_code = secrets.token_hex(5).upper()
    return f"GEEKLOGK-{random_code[:5]}-{random_code[5:]}"


def assign_tracking_code(user_id: int) -> str | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    existing_codes = {
        (user.get("tracking_code", "") or "").upper()
        for user in users.values()
        if user.get("tracking_code")
    }

    tracking_code = generate_tracking_code()
    while tracking_code.upper() in existing_codes:
        tracking_code = generate_tracking_code()

    users[key]["tracking_code"] = tracking_code
    users[key]["tracking_stage"] = TRACKING_STAGES["1"]
    users[key]["last_seen"] = now_str()
    save_users(users)
    return tracking_code


def set_tracking_stage(tracking_code: str, stage_key: str) -> dict | None:
    user_key, user = get_user_by_tracking_code(tracking_code)
    if not user_key or not user:
        return None

    stage = TRACKING_STAGES.get(stage_key)
    if not stage:
        return None

    users = load_users()
    users[user_key]["tracking_stage"] = stage
    users[user_key]["last_seen"] = now_str()
    save_users(users)
    return users[user_key]


def cancel_user_order(user_id: int) -> dict | None:
    users = load_users()
    key = str(user_id)
    if key not in users:
        return None

    users[key]["country"] = ""
    users[key]["status"] = "Новый"
    users[key]["order_number"] = 0
    users[key]["tracking_code"] = ""
    users[key]["tracking_stage"] = ""
    users[key]["last_seen"] = now_str()
    save_users(users)
    return users[key]


def purge_canceled_orders() -> None:
    users = load_users()
    changed = False

    for user in users.values():
        if user.get("status") != "Заказ отменен":
            continue

        user["country"] = ""
        user["status"] = "Новый"
        user["order_number"] = 0
        user["tracking_code"] = ""
        user["tracking_stage"] = ""
        user["last_seen"] = now_str()
        changed = True

    if changed:
        save_users(users)


def is_admin(message: Message) -> bool:
    return bool(message.from_user and message.from_user.id == ADMIN_ID)


async def maintenance_guard(message: Message) -> bool:
    if is_admin(message):
        return False
    if not is_maintenance_mode():
        return False

    await message.answer("Бот временно находится на техобслуживании. Попробуй написать позже.")
    return True


async def banned_guard(message: Message) -> bool:
    if is_admin(message) or not message.from_user:
        return False

    user = get_user(message.from_user.id)
    if not user or int(user.get("is_banned", 0) or 0) != 1:
        return False

    reason = user.get("ban_reason", "") or "без указания причины"
    await message.answer(f"Ты заблокирован в боте.\nПричина: {reason}")
    return True


def is_rate_limited(message: Message) -> bool:
    if not message.from_user:
        return False
    if is_admin(message):
        return False

    user_id = message.from_user.id
    now = time.monotonic()
    history = user_request_times.get(user_id, [])
    history = [stamp for stamp in history if now - stamp <= RATE_WINDOW_SECONDS]
    history.append(now)
    user_request_times[user_id] = history
    return len(history) > MAX_MESSAGES_PER_WINDOW


async def maybe_send_rate_limit_notice(message: Message) -> None:
    if not message.from_user:
        return

    user_id = message.from_user.id
    now = time.monotonic()
    last_notice = user_last_limit_notice.get(user_id, 0)
    if now - last_notice < RATE_LIMIT_NOTICE_COOLDOWN:
        return

    user_last_limit_notice[user_id] = now
    await message.answer("Слишком много сообщений подряд. Подожди пару секунд и отправь снова.")


def current_main_keyboard(message: Message) -> ReplyKeyboardMarkup:
    return admin_main_keyboard if is_admin(message) else user_main_keyboard


def normalize_query(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def tokenize_query(text: str) -> list[str]:
    return [token for token in re.split(r"\W+", normalize_query(text)) if len(token) > 1]


def build_result_text(results: list[dict]) -> str:
    if not results:
        return "Ничего не нашел по этому запросу. Попробуй написать точнее модель, бренд или цвет."

    lines = ["Вот что удалось найти:"]
    for index, result in enumerate(results[:MAX_RESULTS], start=1):
        lines.append(f"{index}. {result['title'][:120]}\n{result['url']}")
    return "\n\n".join(lines)


def split_text_chunks(text: str, limit: int = 3500) -> list[str]:
    if len(text) <= limit:
        return [text]

    chunks: list[str] = []
    current = ""
    for block in text.split("\n\n"):
        candidate = block if not current else f"{current}\n\n{block}"
        if len(candidate) <= limit:
            current = candidate
            continue

        if current:
            chunks.append(current)
            current = ""

        while len(block) > limit:
            chunks.append(block[:limit])
            block = block[limit:]
        current = block

    if current:
        chunks.append(current)
    return chunks


async def send_long_text(message: Message, text: str) -> None:
    for chunk in split_text_chunks(text):
        await message.answer(chunk)


async def apply_tracking_stage_change(
    message: Message,
    bot: Bot,
    tracking_code: str,
    stage_key: str,
) -> bool:
    updated_user = set_tracking_stage(tracking_code, stage_key)
    if not updated_user:
        await message.answer("Заказ с таким кодом не найден.")
        return False

    username = f"@{updated_user['username']}" if updated_user.get("username") else "без username"
    stage = TRACKING_STAGES[stage_key]
    await message.answer(
        f"Этап обновлен.\n"
        f"Код: {tracking_code}\n"
        f"Владелец: {username}\n"
        f"Новый этап: {stage}"
    )
    try:
        await bot.send_message(
            updated_user["id"],
            "Обновление по заказу.\n"
            f"Трек-код: {tracking_code}\n"
            f"Новый этап: {stage}"
        )
    except Exception:
        pass
    return True


async def send_photo_or_text(
    message: Message,
    image_path: Path,
    caption: str,
    reply_markup: ReplyKeyboardMarkup | None = None,
) -> None:
    if image_path.exists():
        await message.answer_photo(
            photo=FSInputFile(str(image_path)),
            caption=caption,
            reply_markup=reply_markup,
        )
        return
    await message.answer(caption, reply_markup=reply_markup)


async def send_main_menu(message: Message, user_name: str) -> None:
    await send_photo_or_text(
        message,
        WELCOME_IMAGE,
        f"Привет, {user_name}! Это бот для заказа вещей из разных стран.\nВыбери нужный раздел в меню ниже:",
        reply_markup=current_main_keyboard(message),
    )


async def notify_admin(bot: Bot, message: Message) -> None:
    if not message.from_user or message.from_user.id == ADMIN_ID:
        return

    user = get_user(message.from_user.id) or {}
    username = f"@{message.from_user.username}" if message.from_user.username else "без username"
    text = message.text or message.caption or "[не текстовое сообщение]"
    tariff_line = ""
    if user.get("tariff"):
        tariff_line = f"Тариф: {user['tariff']}\n"
    await bot.send_message(
        ADMIN_ID,
        "Новые данные по заказу:\n\n"
        f"ID: {message.from_user.id}\n"
        f"Имя: {message.from_user.full_name}\n"
        f"Username: {username}\n"
        f"{tariff_line}"
        f"Текст: {text}\n\n"
        "Чтобы ответить, используй:\n"
        "/send user_id текст\n"
        "/done user_id",
    )


async def fetch_site_html(session: aiohttp.ClientSession, url: str) -> str:
    try:
        async with session.get(url, headers=HTTP_HEADERS, ssl=False) as response:
            if response.status != 200:
                return ""
            return await response.text()
    except Exception:
        return ""


def extract_matches_from_html(html: str, base_url: str, query: str) -> list[dict]:
    if not html:
        return []

    tokens = tokenize_query(query)
    if not tokens:
        return []

    soup = BeautifulSoup(html, "html.parser")
    results: list[dict] = []
    seen_urls: set[str] = set()

    for link in soup.find_all("a", href=True):
        title_parts = [
            link.get_text(" ", strip=True),
            link.get("title", ""),
            link.get("aria-label", ""),
            link.get("alt", ""),
            link["href"],
        ]
        haystack = normalize_query(" ".join(part for part in title_parts if part))
        if not haystack:
            continue

        token_matches = sum(1 for token in tokens if token in haystack)
        if token_matches == 0:
            continue

        href = urljoin(base_url, link["href"])
        if href in seen_urls:
            continue

        seen_urls.add(href)
        title = link.get_text(" ", strip=True) or link.get("title") or href
        results.append({"title": title, "url": href, "score": token_matches})

    results.sort(key=lambda item: item["score"], reverse=True)
    return results[:MAX_RESULTS]


async def search_source(session: aiohttp.ClientSession, source_url: str, query: str) -> list[dict]:
    html = await fetch_site_html(session, source_url)
    return extract_matches_from_html(html, source_url, query)


async def search_item_in_sources(query: str) -> list[dict]:
    connector = aiohttp.TCPConnector(limit=12, limit_per_host=2, ssl=False)
    async with search_semaphore:
        async with aiohttp.ClientSession(timeout=REQUEST_TIMEOUT, connector=connector) as session:
            batches = await asyncio.gather(
                *(search_source(session, url, query) for url in SEARCH_SOURCES),
                return_exceptions=False,
            )

    results = [item for batch in batches for item in batch]
    results.sort(key=lambda item: item["score"], reverse=True)

    unique: list[dict] = []
    seen_urls: set[str] = set()
    for result in results:
        if result["url"] in seen_urls:
            continue
        seen_urls.add(result["url"])
        unique.append(result)
        if len(unique) >= MAX_RESULTS:
            break
    return unique


@dp.message(Command("start"))
async def cmd_start(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if await banned_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message, status="Запустил бота")
    user_name = message.from_user.first_name if message.from_user else "друг"
    await send_main_menu(message, user_name)


@dp.message(Command("help"))
async def cmd_help(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if await banned_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    await message.answer(
        "Команды:\n"
        "/start - запустить бота\n"
        "/help - показать помощь\n"
        "/track КОД - посмотреть статус заказа\n\n"
        "Доступные разделы:\n"
        "- Сделать заказ\n"
        "- Тарифы\n"
        "- Инфо\n"
        "- Тех. поддержка\n\n"
        "В разделе 'Инфо' можно быстро посмотреть свой заказ и попытки поиска."
    )


@dp.message(Command("admin"))
async def admin_panel(message: Message) -> None:
    update_user(message)
    if is_admin(message):
        await message.answer(
            "Админ-панель:\n\n"
            f"Техобслуживание: {'включено' if is_maintenance_mode() else 'выключено'}\n\n"
            "Кнопка 'Оформить заказ' - добавить товар в Excel пошагово\n"
            "/users - список пользователей\n"
            "/orders - все заказы и их статусы\n"
            "/active - только активные заказы\n"
            "/tracks - заказы с трек-кодами\n"
            "/orderno номер - найти заказ по номеру\n"
            "/today - короткая сводка за сегодня\n"
            "/recent - последние 10 пользователей\n"
            "/search запрос - поиск по пользователям и заказам\n"
            "/grantsearch user_id [кол-во] - выдать попытки поиска\n"
            "/setsearch user_id кол-во - выставить точное число попыток\n"
            "/revokesearch user_id - удалить все попытки поиска\n"
            "/searchlist - у кого есть попытки поиска\n"
            "/user user_id - карточка пользователя\n"
            "/owner код - узнать владельца трек-кода\n"
            "/status user_id текст - вручную сменить статус\n"
            "/stats - статистика по боту\n"
            "/dashboard - быстрый дашборд по боту\n"
            "/discount user_id процент - выдать скидку\n"
            "/discountoff user_id - убрать скидку\n"
            "/buyoutfree user_id - включить выкуп без комиссии\n"
            "/buyoutfee user_id - отключить выкуп без комиссии\n"
            "/note user_id текст - заметка по пользователю\n"
            "/clearnote user_id - очистить заметку\n"
            "/ban user_id причина - забанить пользователя\n"
            "/unban user_id - разбанить пользователя\n"
            "/banned - список забаненных пользователей\n"
            "/payreq user_id реквизиты - отправить реквизиты на оплату\n"
            "/paid user_id - отметить оплату как подтвержденную\n"
            "/send user_id текст - отправить сообщение пользователю\n"
            "/broadcast текст - отправить сообщение всем пользователям\n"
            "/maintenance on|off - включить или выключить техобслуживание\n"
            "/setyuan 12,0 - изменить курс юаня\n"
            "/randomuser - случайный пользователь\n"
            "/adminfun - случайная админ-фраза\n"
            "/done user_id - обработать заказ и выдать трек-код\n"
            "/cancel user_id - отменить заказ клиента\n"
            "/track код - посмотреть этап заказа\n"
            "/trackset код этап - изменить этап заказа\n"
            "/vykup код - быстро поставить этап 'Выкуплен'\n"
            "/sklad код - быстро поставить этап 'Приехал на склад'\n"
            "/otpravlen код - быстро поставить этап 'Отправлен'\n"
            "/rf код - быстро поставить этап 'Приехал в РФ'\n"
            "/find текст - поиск вещи по твоим ссылкам\n"
            "/excelsend [номер|имя|current] - отправить таблицу файлом\n"
            "/excelview [номер|имя|current] - показать содержимое таблицы\n"
            "/exceldelete [номер|имя|current] - удалить таблицу\n"
            "/exceledit [номер|имя|current] строка поле значение - изменить строку таблицы\n"
            "/excelorder - добавить товар в текущую таблицу карго\n"
            "/newexcelorder - начать новую таблицу карго\n"
            "/messageorder - оформить заказ и отправить его одним сообщением\n"
            "/excelfiles - показать все таблицы карго\n"
            "/activeexcel - показать текущую активную таблицу\n"
            "/setactiveexcel [номер|имя] - сделать таблицу текущей\n"
            "/closeexcel - закрыть текущую активную таблицу\n"
            "/cargostats - короткая сводка по таблицам карго\n"
            "/admin - показать все админ-команды"
        )
        return
    if not is_admin(message):
        await message.answer("У тебя нет доступа к админ-панели.")
        return

    maintenance_status = "включено" if is_maintenance_mode() else "выключено"
    await message.answer(
        "Админ-панель:\n\n"
        f"Техобслуживание: {maintenance_status}\n\n"
        "/users - список пользователей\n"
        "/orders - все заказы и их статусы\n"
        "/active - только активные заказы\n"
        "/tracks - заказы с трек-кодами\n"
        "/orderno номер - найти заказ по номеру\n"
        "/today - короткая сводка за сегодня\n"
        "/recent - последние 10 пользователей\n"
        "/search запрос - поиск по пользователям и заказам\n"
        "/grantsearch user_id [кол-во] - выдать попытки поиска\n"
        "/setsearch user_id кол-во - выставить точное число попыток\n"
        "/revokesearch user_id - удалить все попытки поиска\n"
        "/searchlist - у кого есть попытки поиска\n"
        "/user user_id - карточка пользователя\n"
        "/owner код - узнать владельца трек-кода\n"
        "/status user_id текст - вручную сменить статус\n"
        "/stats - статистика по боту\n"
        "/note user_id текст - заметка по пользователю\n"
        "/clearnote user_id - очистить заметку\n"
        "/send user_id текст - отправить сообщение пользователю\n"
        "/broadcast текст - отправить сообщение всем пользователям\n"
        "/maintenance on|off - включить или выключить техобслуживание\n"
        "/done user_id - обработать заказ и выдать трек-код\n"
        "/cancel user_id - отменить заказ клиента\n"
        "/track код - посмотреть этап заказа\n"
        "/trackset код этап - изменить этап заказа\n"
        "/vykup код - быстро поставить этап 'Выкуплен'\n"
        "/sklad код - быстро поставить этап 'Приехал на склад'\n"
        "/otpravlen код - быстро поставить этап 'Отправлен'\n"
        "/rf код - быстро поставить этап 'Приехал в РФ'\n"
        "/find текст - поиск вещи по твоим ссылкам\n\n"
        "Поиск для клиентов:\n"
        "Поиск по фото доступен только админу.\n"
        "Обычным пользователям можно выдать платные попытки поиска по названию.\n\n"
        "Этапы доставки:\n"
        "1 - Выкуплен\n"
        "2 - Приехал на склад\n"
        "3 - Отправлен\n"
        "4 - Приехал в РФ"
    )


@dp.message(Command("stats"))
async def admin_stats(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = load_users()
    total = len(users)
    orders = sum(1 for user in users.values() if user.get("order_number", 0))
    china = sum(1 for user in users.values() if user.get("country") == "Китай")
    sent = sum(1 for user in users.values() if user.get("status") == "Заявка отправлена")
    pending = sum(1 for user in users.values() if user.get("status") == "Отправляет данные")
    processed = sum(1 for user in users.values() if user.get("status") == "Заказ обработан")
    with_codes = sum(1 for user in users.values() if user.get("tracking_code"))
    paid_search = sum(1 for user in users.values() if int(user.get("search_access", 0) or 0) > 0)
    total_search_credits = sum(int(user.get("search_access", 0) or 0) for user in users.values())
    in_russia = sum(1 for user in users.values() if user.get("tracking_stage") == TRACKING_STAGES["4"])

    await message.answer(
        "Статистика:\n\n"
        f"Всего пользователей: {total}\n"
        f"Всего заказов: {orders}\n"
        f"Китай: {china}\n"
        f"Отправляют данные: {pending}\n"
        f"Заявка отправлена: {sent}\n"
        f"Заказ обработан: {processed}\n"
        f"С трек-кодом: {with_codes}\n"
        f"Пользователей с поиском: {paid_search}\n"
        f"Всего попыток поиска: {total_search_credits}\n"
        f"Приехали в РФ: {in_russia}"
    )


@dp.message(Command("maintenance"))
async def admin_maintenance(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        status = "включено" if is_maintenance_mode() else "выключено"
        await message.answer(
            "Используй формат: /maintenance on|off\n"
            f"Сейчас техобслуживание: {status}"
        )
        return

    mode = parts[1].strip().lower()
    if mode not in {"on", "off"}:
        await message.answer("Используй формат: /maintenance on|off")
        return

    enabled = mode == "on"
    set_maintenance_mode(enabled)
    await message.answer(
        "Техобслуживание включено. Теперь бот отвечает только админу."
        if enabled
        else "Техобслуживание выключено. Бот снова отвечает всем пользователям."
    )


@dp.message(Command("broadcast"))
async def admin_broadcast(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /broadcast текст")
        return

    users = load_users()
    recipients = [
        user["id"]
        for user in users.values()
        if int(user.get("id", 0) or 0) != ADMIN_ID
    ]
    if not recipients:
        await message.answer("Нет пользователей для рассылки.")
        return

    text = parts[1].strip()
    sent = 0
    failed = 0

    await message.answer(f"Начинаю рассылку. Получателей: {len(recipients)}")
    for user_id in recipients:
        try:
            await bot.send_message(user_id, text)
            sent += 1
        except Exception:
            failed += 1

    await message.answer(
        "Рассылка завершена.\n"
        f"Отправлено: {sent}\n"
        f"Не доставлено: {failed}"
    )


@dp.message(Command("pending"))
async def pending_orders(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    pending = get_pending_orders()
    if not pending:
        await message.answer("Неодобренных заказов сейчас нет.")
        return

    lines = [format_order_card(user, include_action_hint=True) for user in pending]
    await send_long_text(message, "Неодобренные заказы:\n\n" + "\n\n".join(lines))


@dp.message(Command("orders"))
async def all_orders(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    orders = get_all_orders()
    if not orders:
        await message.answer("Заказов пока нет.")
        return

    lines = [format_order_card(user, include_action_hint=True) for user in orders]
    await send_long_text(message, "Все заказы и их статусы:\n\n" + "\n\n".join(lines))


@dp.message(Command("active"))
async def active_orders(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    orders = get_active_orders()
    if not orders:
        await message.answer("Активных заказов сейчас нет.")
        return

    lines = [format_order_card(user, include_action_hint=True) for user in orders]
    await send_long_text(message, "Активные заказы:\n\n" + "\n\n".join(lines))


@dp.message(Command("tracks"))
async def tracked_orders(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    orders = get_tracked_orders()
    if not orders:
        await message.answer("Заказов с трек-кодами пока нет.")
        return

    lines = [format_order_card(user) for user in orders]
    await send_long_text(message, "Заказы с трек-кодами:\n\n" + "\n\n".join(lines))


@dp.message(Command("orderno"))
async def order_by_number(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /orderno номер")
        return

    try:
        order_number = int(parts[1].strip())
    except ValueError:
        await message.answer("Номер заказа должен быть числом.")
        return

    user = get_user_by_order_number(order_number)
    if not user:
        await message.answer("Заказ с таким номером не найден.")
        return

    await message.answer("Заказ найден:\n\n" + format_order_card(user, include_action_hint=True))


@dp.message(Command("today"))
async def today_summary(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    stats = get_today_stats()
    await message.answer(
        "Сегодня:\n\n"
        f"Новых пользователей: {stats['new_users']}\n"
        f"Активных сегодня: {stats['active_today']}\n"
        f"Заказов обновлялось сегодня: {stats['orders_today']}\n"
        f"Обработано сегодня: {stats['processed_today']}"
    )


@dp.message(Command("recent"))
async def recent_users(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    recent = get_recent_users()
    if not recent:
        await message.answer("Пока нет пользователей.")
        return

    lines = [format_order_card(user) for user in recent]
    await send_long_text(message, "Последние 10 пользователей:\n\n" + "\n\n".join(lines))


@dp.message(Command("search"))
async def admin_search_command(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /search запрос")
        return

    results = search_users(parts[1].strip())
    if not results:
        await message.answer("Ничего не найдено.")
        return

    lines = [format_order_card(user) for user in results]
    await send_long_text(message, "Результаты поиска:\n\n" + "\n\n".join(lines))


@dp.message(Command("grantsearch"))
async def admin_grant_search(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 2:
        await message.answer("Используй формат: /grantsearch user_id [кол-во]")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    amount = 1
    if len(parts) >= 3 and parts[2].strip():
        try:
            amount = int(parts[2].strip())
        except ValueError:
            await message.answer("Количество должно быть числом.")
            return
    if amount <= 0:
        await message.answer("Количество попыток должно быть больше нуля.")
        return

    total_credits = add_search_credits(user_id, amount)
    if total_credits is None:
        await message.answer("Пользователь не найден. Сначала он должен написать боту.")
        return

    try:
        await bot.send_message(
            user_id,
            "Тебе открыт доступ к платному поиску вещи в Китае.\n"
            f"Выдано попыток: {amount}\n"
            f"Всего попыток осталось: {total_credits}\n"
            "Теперь можно зайти в 'Сделать заказ' -> 'Найти эту вещь в Китае'.\n"
            "Поиск по фото доступен только админу, тебе доступен поиск по названию вещи."
        )
    except Exception:
        pass

    await message.answer(
        f"Пользователю {user_id} выдано попыток: {amount}\n"
        f"Всего попыток у него: {total_credits}"
    )


@dp.message(Command("setsearch"))
async def admin_set_search(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /setsearch user_id кол-во")
        return

    try:
        user_id = int(parts[1].strip())
        amount = int(parts[2].strip())
    except ValueError:
        await message.answer("user_id и количество должны быть числами.")
        return

    if amount < 0:
        await message.answer("Количество попыток не может быть отрицательным.")
        return

    remaining = set_search_credits(user_id, amount)
    if remaining is None:
        await message.answer("Пользователь не найден. Сначала он должен написать боту.")
        return

    if remaining <= 0:
        paid_search_users.discard(user_id)

    try:
        await bot.send_message(
            user_id,
            "Обновлен доступ к поиску вещи в Китае.\n"
            f"Сейчас попыток поиска: {remaining}"
        )
    except Exception:
        pass

    await message.answer(f"Пользователю {user_id} выставлено попыток поиска: {remaining}")


@dp.message(Command("revokesearch"))
async def admin_revoke_search(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /revokesearch user_id")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    remaining = set_search_credits(user_id, 0)
    if remaining is None:
        await message.answer("Пользователь не найден.")
        return

    paid_search_users.discard(user_id)
    try:
        await bot.send_message(
            user_id,
            "Все попытки поиска вещи в Китае были удалены."
        )
    except Exception:
        pass

    await message.answer(f"Все попытки поиска удалены у пользователя {user_id}.")


@dp.message(Command("searchlist"))
async def admin_search_list(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = get_search_access_users()
    if not users:
        await message.answer("Сейчас ни у кого нет доступа к платному поиску.")
        return

    lines = [format_order_card(user) for user in users]
    await send_long_text(message, "Платные попытки поиска есть у:\n\n" + "\n\n".join(lines))


@dp.message(Command("user"))
async def admin_user_card(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /user user_id")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = get_user(user_id)
    if not user:
        await message.answer("Пользователь не найден.")
        return

    username = f"@{user['username']}" if user.get("username") else "без username"
    full_name = " ".join(
        part for part in [user.get("first_name", ""), user.get("last_name", "")] if part
    ).strip() or "Без имени"

    await message.answer(
        "Карточка пользователя:\n\n"
        f"Номер заказа: {user.get('order_number', 0) or 'Нет'}\n"
        f"ID: {user['id']}\n"
        f"Имя: {full_name}\n"
        f"Username: {username}\n"
        f"Дата первого входа: {user.get('first_seen', '-')}\n"
        f"Последняя активность: {user.get('last_seen', '-')}\n"
        f"Страна: {user.get('country', '') or 'Не выбрана'}\n"
        f"Тариф: {user.get('tariff', '') or 'Не выбран'}\n"
        f"Статус: {user.get('status', '') or 'Новый'}\n"
        f"Скидка: {int(user.get('discount', 0) or 0)}%\n"
        f"Выкуп без комиссии: {'Да' if int(user.get('buyout_free', 0) or 0) == 1 else 'Нет'}\n"
        f"Забанен: {'Да' if int(user.get('is_banned', 0) or 0) == 1 else 'Нет'}\n"
        f"Причина бана: {user.get('ban_reason', '') or 'Нет'}\n"
        f"Оплата: {user.get('payment_status', '') or 'Нет'}\n"
        f"Попыток поиска: {int(user.get('search_access', 0) or 0)}\n"
        f"Трек-код: {user.get('tracking_code', '') or 'Нет'}\n"
        f"Этап доставки: {user.get('tracking_stage', '') or 'Нет'}\n"
        f"Заметка: {user.get('note', '') or 'Нет'}"
    )


@dp.message(Command("owner"))
async def tracking_owner(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /owner код")
        return

    tracking_code = parts[1].strip().upper()
    _, user = get_user_by_tracking_code(tracking_code)
    if not user:
        await message.answer("Заказ с таким кодом не найден.")
        return

    await message.answer(
        "Владелец трек-кода:\n\n"
        f"{format_order_card(user)}"
    )


@dp.message(Command("note"))
async def admin_note(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /note user_id текст")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    note = parts[2].strip()
    if not note:
        await message.answer("Заметка не должна быть пустой.")
        return

    if not set_user_note(user_id, note):
        await message.answer("Пользователь не найден.")
        return

    await message.answer(f"Заметка для пользователя {user_id} сохранена.")


@dp.message(Command("clearnote"))
async def admin_clear_note(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /clearnote user_id")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    if not set_user_note(user_id, ""):
        await message.answer("Пользователь не найден.")
        return

    await message.answer(f"Заметка пользователя {user_id} очищена.")


@dp.message(Command("status"))
async def admin_status(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /status user_id новый_статус")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = get_user(user_id)
    if not user:
        await message.answer("Пользователь не найден.")
        return

    new_status = parts[2].strip()
    if not new_status:
        await message.answer("Статус не должен быть пустым.")
        return

    set_user_status(user_id, new_status)
    await message.answer(f"Статус пользователя {user_id} обновлен: {new_status}")


@dp.message(Command("find"))
async def admin_find(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /find название вещи")
        return

    query = parts[1].strip()
    await message.answer("Ищу по твоим ссылкам, подожди немного...")
    results = await search_item_in_sources(query)
    await message.answer(build_result_text(results))


@dp.message(Command("track"))
async def track_order(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /track код")
        return

    tracking_code = parts[1].strip().upper()
    _, user = get_user_by_tracking_code(tracking_code)
    if not user:
        await message.answer("Заказ с таким кодом не найден.")
        return

    await message.answer(
        f"Трек-код: {tracking_code}\n"
        f"Текущий этап: {user.get('tracking_stage', '') or 'Этап пока не назначен'}"
    )


@dp.message(Command("trackset"))
async def track_set(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /trackset код этап")
        return

    tracking_code = parts[1].strip().upper()
    stage_key = parts[2].strip()
    if stage_key not in TRACKING_STAGES:
        await message.answer(
            "Доступные этапы:\n"
            "1 - Выкуплен\n"
            "2 - Приехал на склад\n"
            "3 - Отправлен\n"
            "4 - Приехал в РФ"
        )
        return

    await apply_tracking_stage_change(message, bot, tracking_code, stage_key)


async def handle_quick_stage(message: Message, bot: Bot, stage_key: str) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Отправь команду с трек-кодом.")
        return

    tracking_code = parts[1].strip().upper()
    await apply_tracking_stage_change(message, bot, tracking_code, stage_key)


@dp.message(Command("vykup"))
async def quick_vykup(message: Message, bot: Bot) -> None:
    await handle_quick_stage(message, bot, "1")


@dp.message(Command("sklad"))
async def quick_sklad(message: Message, bot: Bot) -> None:
    await handle_quick_stage(message, bot, "2")


@dp.message(Command("otpravlen"))
async def quick_otpravlen(message: Message, bot: Bot) -> None:
    await handle_quick_stage(message, bot, "3")


@dp.message(Command("rf"))
async def quick_rf(message: Message, bot: Bot) -> None:
    await handle_quick_stage(message, bot, "4")


@dp.message(Command("users"))
async def list_users(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = load_users()
    if not users:
        await message.answer("Пока нет пользователей, которые писали боту.")
        return

    lines = []
    for user in sorted(users.values(), key=lambda item: item.get("last_seen", ""), reverse=True):
        username = f"@{user['username']}" if user.get("username") else "без username"
        full_name = " ".join(
            part for part in [user.get("first_name", ""), user.get("last_name", "")] if part
        ).strip() or "Без имени"
        lines.append(
            f"Заказ №{user.get('order_number', 0) or '-'} | {user['id']} | {full_name} | {username}\n"
            f"Дата: {user.get('first_seen', '-')}\n"
            f"Страна: {user.get('country', '') or 'Не выбрана'}\n"
            f"Статус: {user.get('status', '') or 'Новый'}\n"
            f"Попыток поиска: {int(user.get('search_access', 0) or 0)}\n"
            f"Трек-код: {user.get('tracking_code', '') or 'Нет'}\n"
            f"Этап: {user.get('tracking_stage', '') or 'Нет'}"
        )

    await send_long_text(message, "Пользователи:\n\n" + "\n\n".join(lines))


@dp.message(Command("send"))
async def send_to_user(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /send user_id текст")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    try:
        await bot.send_message(user_id, parts[2].strip())
        await message.answer(f"Сообщение отправлено пользователю {user_id}.")
    except Exception as error:
        await message.answer(f"Не удалось отправить сообщение: {error}")


@dp.message(Command("done"))
async def send_done_message(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /done user_id")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = get_user(user_id)
    if not can_approve_order(user):
        await message.answer(
            "Нельзя одобрить этот заказ. Код выдается только если пользователь уже отправил заявку."
        )
        return

    try:
        tracking_code = assign_tracking_code(user_id)
        await bot.send_message(
            user_id,
            "Ваш заказ подтвержден.\n"
            "Теперь можно перейти к оплате.\n"
            "Нажми кнопку 'Оплатить заказ', чтобы запросить реквизиты.",
            reply_markup=payment_keyboard,
        )
        update_user_fields(
            user_id,
            status="Ожидает оплату",
            payment_status="approved_waiting_payment",
            tracking_code=tracking_code,
            tracking_stage=TRACKING_STAGES["1"],
        )
        await message.answer(
            "Заказ подтвержден и переведен в оплату.\n"
            f"Пользователь: {user_id}\n"
            f"Трек-код уже создан: {tracking_code}"
        )
    except Exception as error:
        await message.answer(f"Не удалось отправить сообщение: {error}")


@dp.message(Command("cancel"))
async def cancel_order(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /cancel user_id")
        return

    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = get_user(user_id)
    if not can_cancel_order(user):
        await message.answer("Этот заказ нельзя отменить или пользователь не найден.")
        return

    order_number = user.get("order_number", 0) if user else 0
    tracking_code = user.get("tracking_code", "") if user else ""
    canceled_user = cancel_user_order(user_id)
    china_order_users.discard(user_id)
    admin_search_users.discard(user_id)
    tracking_lookup_users.discard(user_id)

    if not canceled_user:
        await message.answer("Не удалось отменить заказ.")
        return

    try:
        lines = ["Ваш заказ отменен."]
        if order_number:
            lines.append(f"Номер заказа: {order_number}")
        if tracking_code:
            lines.append(f"Трек-код: {tracking_code}")

        await bot.send_message(
            user_id,
            "\n".join(lines),
            reply_markup=user_main_keyboard,
        )
    except Exception:
        pass

    await message.answer(
        "Заказ отменен.\n"
        f"Пользователь: {user_id}\n"
        f"Номер заказа: {order_number or 'Нет'}\n"
        f"Трек-код: {tracking_code or 'Нет'}"
    )


@dp.message(Command("dashboard"))
async def admin_dashboard(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = load_users()
    await message.answer(
        "Дашборд:\n\n"
        f"Пользователей: {len(users)}\n"
        f"Активных заказов: {sum(1 for user in users.values() if user.get('status') in ACTIVE_ORDER_STATUSES)}\n"
        f"Забаненных: {sum(1 for user in users.values() if int(user.get('is_banned', 0) or 0) == 1)}\n"
        f"Со скидкой: {sum(1 for user in users.values() if int(user.get('discount', 0) or 0) > 0)}\n"
        f"Выкуп без комиссии: {sum(1 for user in users.values() if int(user.get('buyout_free', 0) or 0) == 1)}\n"
        f"Ожидают оплату: {sum(1 for user in users.values() if (user.get('payment_status', '') or '') == 'waiting_payment')}\n"
        f"Оплачено: {sum(1 for user in users.values() if (user.get('payment_status', '') or '') == 'paid')}"
    )


@dp.message(Command("discount"))
async def admin_discount(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /discount user_id процент")
        return
    try:
        user_id = int(parts[1].strip())
        percent = int(parts[2].strip())
    except ValueError:
        await message.answer("user_id и процент должны быть числами.")
        return
    if percent < 0 or percent > 100:
        await message.answer("Процент скидки должен быть от 0 до 100.")
        return

    user = update_user_fields(user_id, discount=percent)
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(user_id, f"Тебе выдана скидка {percent}%.")
    except Exception:
        pass
    await message.answer(f"Пользователю {user_id} выдана скидка {percent}%.")


@dp.message(Command("discountoff"))
async def admin_discount_off(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /discountoff user_id")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = update_user_fields(user_id, discount=0)
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(user_id, "Твоя скидка отключена.")
    except Exception:
        pass
    await message.answer(f"Скидка у пользователя {user_id} отключена.")


@dp.message(Command("buyoutfree"))
async def admin_buyout_free(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /buyoutfree user_id")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = update_user_fields(user_id, buyout_free=1)
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(user_id, "Тебе включен выкуп без комиссии.")
    except Exception:
        pass
    await message.answer(f"Пользователю {user_id} включен выкуп без комиссии.")


@dp.message(Command("buyoutfee"))
async def admin_buyout_fee(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /buyoutfee user_id")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = update_user_fields(user_id, buyout_free=0)
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(user_id, "Выкуп без комиссии отключен.")
    except Exception:
        pass
    await message.answer(f"У пользователя {user_id} отключен выкуп без комиссии.")


@dp.message(Command("ban"))
async def admin_ban(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /ban user_id причина")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    reason = parts[2].strip()
    if not reason:
        await message.answer("Причина бана не должна быть пустой.")
        return

    user = update_user_fields(user_id, is_banned=1, ban_reason=reason)
    if not user:
        await message.answer("Пользователь не найден.")
        return
    paid_search_users.discard(user_id)
    admin_search_users.discard(user_id)
    tracking_lookup_users.discard(user_id)
    try:
        await bot.send_message(user_id, f"Ты заблокирован в боте.\nПричина: {reason}")
    except Exception:
        pass
    await message.answer(f"Пользователь {user_id} забанен.\nПричина: {reason}")


@dp.message(Command("unban"))
async def admin_unban(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /unban user_id")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = update_user_fields(user_id, is_banned=0, ban_reason="")
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(user_id, "Ты разблокирован в боте.")
    except Exception:
        pass
    await message.answer(f"Пользователь {user_id} разбанен.")


@dp.message(Command("banned"))
async def admin_banned(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = get_banned_users()
    if not users:
        await message.answer("Список забаненных пуст.")
        return
    lines = [
        f"{user['id']} | {' '.join(part for part in [user.get('first_name', ''), user.get('last_name', '')] if part).strip() or 'Без имени'}\n"
        f"Причина: {user.get('ban_reason', '') or 'не указана'}"
        for user in users
    ]
    await send_long_text(message, "Забаненные пользователи:\n\n" + "\n\n".join(lines))


@dp.message(Command("payreq"))
async def admin_payreq(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Используй формат: /payreq user_id реквизиты")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    requisites = parts[2].strip()
    if not requisites:
        await message.answer("Реквизиты не должны быть пустыми.")
        return

    user = update_user_fields(user_id, payment_status="waiting_payment")
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        await bot.send_message(
            user_id,
            "Реквизиты для оплаты:\n"
            f"{requisites}\n\n"
            "После оплаты отправь чек или подтверждение в бот."
        )
    except Exception as error:
        await message.answer(f"Не удалось отправить реквизиты: {error}")
        return
    await message.answer(f"Реквизиты отправлены пользователю {user_id}.")


@dp.message(Command("paid"))
async def admin_paid(message: Message, bot: Bot) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Используй формат: /paid user_id")
        return
    try:
        user_id = int(parts[1].strip())
    except ValueError:
        await message.answer("user_id должен быть числом.")
        return

    user = update_user_fields(user_id, payment_status="paid", status="Заказ обработан")
    if not user:
        await message.answer("Пользователь не найден.")
        return
    try:
        tracking_code = user.get("tracking_code", "") or "пока не назначен"
        await bot.send_message(
            user_id,
            "Оплата подтверждена. Спасибо!\n"
            f"Твой трек-код: {tracking_code}\n"
            f"Текущий этап: {user.get('tracking_stage', '') or TRACKING_STAGES['1']}",
            reply_markup=current_main_keyboard(message),
        )
    except Exception:
        pass
    await message.answer(f"Оплата пользователя {user_id} подтверждена.")


@dp.message(Command("randomuser"))
async def admin_random_user(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    users = list(load_users().values())
    if not users:
        await message.answer("Пока нет пользователей.")
        return
    await message.answer("Случайный пользователь:\n\n" + format_order_card(secrets.choice(users)))


@dp.message(Command("adminfun"))
async def admin_fun(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    phrases = [
        "Админ в потоке. Бот под контролем.",
        "Все спокойно, заказы дышат ровно.",
        "Панель управления сегодня особенно послушная.",
        "Склад мысленно уже собран и отправлен.",
        "Работаем красиво и без суеты.",
    ]
    await message.answer(secrets.choice(phrases))


@dp.message(Command("excelorder"))
async def admin_excel_order(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    ensure_cargo_dirs()
    file_path = get_active_cargo_file() or create_cargo_file()
    set_active_cargo_file(file_path)
    admin_excel_sessions[message.from_user.id] = {
        "mode": "excel_order",
        "step": "photo",
        "file_path": file_path,
        "data": {},
    }
    await message.answer(
        f"Добавляем товар в таблицу {file_path.name}.\n"
        "Отправь фото товара."
    )


@dp.message(Command("newexcelorder"))
async def admin_new_excel_order(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    ensure_cargo_dirs()
    file_path = create_cargo_file()
    admin_excel_sessions[message.from_user.id] = {
        "mode": "excel_order",
        "step": "photo",
        "file_path": file_path,
        "data": {},
    }
    await message.answer(
        f"Создана новая таблица {file_path.name}.\n"
        "Отправь фото товара."
    )


@dp.message(Command("excelfiles"))
async def admin_excel_files(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    files = list_cargo_files()
    if not files:
        await message.answer("Пока нет таблиц карго.")
        return
    active = get_active_cargo_file()
    lines = []
    for index, file_path in enumerate(files, start=1):
        marker = " [active]" if active and active == file_path else ""
        lines.append(f"{index}. {file_path.name}{marker}")
    await message.answer("Таблицы карго:\n\n" + "\n".join(lines))


@dp.message(Command("activeexcel"))
async def admin_active_excel(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    active = get_active_cargo_file()
    if not active:
        await message.answer("Сейчас нет активной таблицы.")
        return
    await message.answer(f"Текущая активная таблица: {active.name}")


@dp.message(Command("excelsend"))
async def admin_excel_send(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    parts = (message.text or "").split(maxsplit=1)
    file_path = resolve_cargo_file(parts[1] if len(parts) > 1 else "current")
    if not file_path:
        await message.answer("Таблица не найдена.")
        return
    await message.answer_document(FSInputFile(str(file_path)))


@dp.message(Command("excelview"))
async def admin_excel_view(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    parts = (message.text or "").split(maxsplit=1)
    file_path = resolve_cargo_file(parts[1] if len(parts) > 1 else "current")
    if not file_path:
        await message.answer("Таблица не найдена.")
        return
    rows = read_cargo_rows(file_path)
    if not rows:
        await message.answer(f"В таблице {file_path.name} пока нет товаров.")
        return
    lines = [f"Таблица: {file_path.name}"]
    for item in rows:
        lines.append(
            f"Строка {item['row']}: {item['name']}\n"
            f"Ссылка: {item['link'] or '-'}\n"
            f"Характеристики: {item['specs'] or '-'}\n"
            f"Цена: {item['price'] or '-'} | Кол-во: {item['quantity'] or '-'} | Итого: {item['total'] or '-'}\n"
            f"Доставка: {item['delivery'] or '-'}"
        )
    await send_long_text(message, "\n\n".join(lines))


@dp.message(Command("exceldelete"))
async def admin_excel_delete(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    parts = (message.text or "").split(maxsplit=1)
    file_path = resolve_cargo_file(parts[1] if len(parts) > 1 else "current")
    if not file_path:
        await message.answer("Таблица не найдена.")
        return
    if get_active_cargo_file() == file_path:
        set_active_cargo_file(None)
    file_path.unlink(missing_ok=True)
    await message.answer(f"Таблица {file_path.name} удалена.")


@dp.message(Command("exceledit"))
async def admin_excel_edit(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return
    parts = (message.text or "").split(maxsplit=4)
    if len(parts) < 5:
        await message.answer("Используй формат: /exceledit [номер|имя|current] строка поле значение")
        return
    file_path = resolve_cargo_file(parts[1])
    if not file_path:
        await message.answer("Таблица не найдена.")
        return
    try:
        row = int(parts[2].strip())
    except ValueError:
        await message.answer("Строка должна быть числом.")
        return
    field = parts[3].strip().lower()
    value = parts[4].strip()
    column_map = {
        "name": "C",
        "link": "D",
        "specs": "E",
        "price": "F",
        "quantity": "G",
        "total": "H",
        "delivery": "I",
    }
    column = column_map.get(field)
    if not column:
        await message.answer("Поле должно быть одним из: name, link, specs, price, quantity, total, delivery")
        return
    workbook = load_workbook(file_path)
    sheet = workbook.active
    sheet[f"{column}{row}"] = value
    workbook.save(file_path)
    await message.answer(f"Таблица {file_path.name}: строка {row}, поле {field} обновлено.")


@dp.message(F.text == "Неодобренные заказы")
async def pending_orders_button(message: Message) -> None:
    if not is_admin(message):
        return

    pending = get_pending_orders()
    if not pending:
        await message.answer("Неодобренных заказов сейчас нет.")
        return

    lines = [format_order_card(user, include_action_hint=True) for user in pending]
    await send_long_text(message, "Неодобренные заказы:\n\n" + "\n\n".join(lines))


@dp.message(F.text == "Все заказы и их статусы")
async def all_orders_button(message: Message) -> None:
    if not is_admin(message):
        return

    orders = get_all_orders()
    if not orders:
        await message.answer("Заказов пока нет.")
        return

    lines = [format_order_card(user, include_action_hint=True) for user in orders]
    await send_long_text(message, "Все заказы и их статусы:\n\n" + "\n\n".join(lines))


@dp.message(F.text == "Сделать заказ")
async def make_order(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if await banned_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message, status="Выбирает раздел заказа")
    await send_photo_or_text(
        message,
        ORDER_IMAGE,
        "Выбери страну, из которой хочешь сделать заказ:",
        reply_markup=user_order_keyboard,
    )


@dp.message(F.text == "Найти эту вещь в Китае")
async def search_item_tab(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return

    user = update_user(message)
    if not message.from_user or not user:
        return

    user_id = message.from_user.id
    china_order_users.discard(user_id)
    tracking_lookup_users.discard(user_id)

    if is_admin(message):
        paid_search_users.discard(user_id)
        admin_search_users.add(user_id)
        await message.answer(
            "Пришли название вещи текстом или фото с подписью для поиска по твоим ссылкам.\n"
            "Чтобы выйти из режима поиска, нажми 'Назад'."
        )
        return

    admin_search_users.discard(user_id)
    credits = int(user.get("search_access", 0) or 0)
    if credits <= 0:
        await message.answer(
            "Поиск вещи в Китае для пользователей доступен только после оплаты.\n"
            "Чтобы получить доступ, напиши в техподдержку: @ichov",
            reply_markup=current_main_keyboard(message),
        )
        return

    paid_search_users.add(user_id)
    await message.answer(
        "Доступ к поиску активирован.\n"
        f"Осталось попыток поиска: {credits}\n"
        "Пришли название вещи текстом, и я начну поиск.\n"
        "Поиск по фото доступен только админу.\n"
        "Чтобы выйти из режима поиска, нажми 'Назад'."
    )


@dp.message(F.text == "Тарифы")
async def tariffs(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    await send_photo_or_text(
        message,
        TARIFF_IMAGE,
        "НАШ ТАРИФ, что он включает?!\n"
        "он включает два вида доставки\n\n"
        "1.Китай-Благовещенск-Керчь\n"
        "2.Китай-Москва-Керчь\n\n"
        "Ценник фиксированный - 6$ за кг\n\n"
        "ДОСТАВКА СДЭК ОПЛАЧИВАЕТСЯ ОТДЕЛЬНО",
    )
    return
    await send_photo_or_text(
        message,
        TARIFF_IMAGE,
        "Пока что существует 1 единственный тариф, это 20-30 дней "
        "с момента отправки со склада - 6 долларов за кг",
    )


@dp.message(Command("setyuan"))
async def set_yuan_rate_command(message: Message) -> None:
    if not is_admin(message):
        await message.answer("У тебя нет доступа к этой команде.")
        return

    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await message.answer("Используй формат: /setyuan 12,0")
        return

    raw_input = parts[1].strip()
    numeric_value = raw_input.replace(",", ".")
    try:
        float(numeric_value)
    except ValueError:
        await message.answer("Курс должен быть числом, например 12,0 или 12.0")
        return

    display_value = raw_input.replace(".", ",")
    set_yuan_rate(display_value)
    await message.answer(f"Курс юаня обновлен: {display_value}")


@dp.message(F.text == "Курс юаня")
async def yuan_rate_tab(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    await send_photo_or_text(
        message,
        YUAN_IMAGE,
        f"Курс юаня {get_yuan_rate()}",
    )


@dp.message(F.text == "Инфо")
async def info(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    tracking_lookup_users.discard(message.from_user.id if message.from_user else -1)
    await send_photo_or_text(
        message,
        INFO_IMAGE,
        "Фотоотчет отправляется прямо в бот.\n"
        "Время работы бота: 8:00 - 20:00 (временно).\n"
        "Здесь ты можешь узнать статус заказа, посмотреть свой заказ и остаток попыток поиска.",
        reply_markup=info_keyboard,
    )


@dp.message(F.text == "Мой заказ")
async def my_order(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    if not message.from_user:
        return

    user = get_user(message.from_user.id)
    if not user:
        await message.answer("Пока нет данных по твоему заказу.", reply_markup=info_keyboard)
        return

    has_order = bool(
        user.get("order_number", 0)
        or user.get("tracking_code")
        or user.get("status") in ACTIVE_ORDER_STATUSES
    )
    if not has_order:
        await message.answer(
            "У тебя пока нет активного заказа.\n"
            "Если хочешь оформить заказ, зайди в раздел 'Сделать заказ'.",
            reply_markup=info_keyboard,
        )
        return

    await message.answer(
        "Твой заказ:\n\n"
        f"Номер заказа: {user.get('order_number', 0) or 'Еще не присвоен'}\n"
        f"Тариф: {user.get('tariff', '') or 'Не выбран'}\n"
        f"Статус: {user.get('status', '') or 'Новый'}\n"
        f"Трек-код: {user.get('tracking_code', '') or 'Пока нет'}\n"
        f"Этап: {user.get('tracking_stage', '') or 'Пока не назначен'}",
        reply_markup=info_keyboard,
    )


@dp.message(F.text == "Мои попытки поиска")
async def my_search_attempts(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    if not message.from_user:
        return

    if is_admin(message):
        await message.answer(
            "У тебя как у админа поиск по фото и по названию доступен без ограничений.",
            reply_markup=info_keyboard,
        )
        return

    user = get_user(message.from_user.id)
    credits = int((user or {}).get("search_access", 0) or 0)
    if credits <= 0:
        await message.answer(
            "У тебя сейчас нет попыток поиска.\n"
            "Чтобы получить платный доступ, напиши в техподдержку: @ichov",
            reply_markup=info_keyboard,
        )
        return

    await message.answer(
        f"У тебя осталось попыток поиска: {credits}\n"
        "Использовать их можно в разделе 'Сделать заказ' -> 'Найти эту вещь в Китае'.",
        reply_markup=info_keyboard,
    )


@dp.message(F.text == "Узнать статус заказа")
async def ask_tracking_code(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    if message.from_user:
        tracking_lookup_users.add(message.from_user.id)
    await message.answer(
        "Отправь свой трек-код, и я покажу текущий статус заказа.",
        reply_markup=info_keyboard,
    )


@dp.message(F.text == "Тех. поддержка")
async def support(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message)
    await send_photo_or_text(
        message,
        SUPPORT_IMAGE,
        "По всем вопросам техподдержки пиши: @ichov",
    )


@dp.message(F.text == "Китай")
async def china_order(message: Message, bot: Bot) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    user = update_user(message, country="Китай", tariff="", status="Выбирает тариф")
    if not message.from_user or not user:
        return

    china_order_users.discard(message.from_user.id)
    tracking_lookup_users.discard(message.from_user.id)
    await message.answer(
        "Выбери тариф доставки:",
        reply_markup=china_tariff_keyboard,
    )


@dp.message(F.text.in_([
    "Китай-Благовещенск-Керчь (2-5 дней до РФ, Владивостока)",
    "Китай-Москва-Керчь (20-30 дней до РФ, МСК)",
]))
async def select_china_tariff(message: Message, bot: Bot) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    if not message.from_user:
        return

    user = update_user(
        message,
        country="Китай",
        tariff=message.text,
        status="Оформляет заказ",
    )
    if not user:
        return

    china_order_users.add(message.from_user.id)
    tracking_lookup_users.discard(message.from_user.id)
    await message.answer(
        "Теперь скидывай ссылку товара, фото товара, размер, цвет и остальные данные по заказу.",
        reply_markup=china_submit_keyboard,
    )
    await bot.send_message(
        ADMIN_ID,
        f"Пользователь {message.from_user.full_name} "
        f"(ID: {message.from_user.id}) начал заказ из Китая.\n"
        f"Тариф: {message.text}",
    )


@dp.message(F.text == "Я все скинул")
async def finish_china_order(message: Message, bot: Bot) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    if not message.from_user or message.from_user.id not in china_order_users:
        update_user(message)
        await message.answer("Сначала выбери раздел 'Сделать заказ', затем 'Китай' и тариф доставки.")
        return

    order_number = assign_order_number(message.from_user.id)
    user = update_user(message, country="Китай", status="Заявка отправлена")
    china_order_users.discard(message.from_user.id)
    tariff = (user or {}).get("tariff", "") or "Не выбран"
    await message.answer(
        f"Готово, ожидайте. Ваш номер заказа: {order_number}",
        reply_markup=current_main_keyboard(message),
    )
    await bot.send_message(
        ADMIN_ID,
        f"Заказ №{order_number}\n"
        f"Пользователь {message.from_user.full_name} "
        f"(ID: {message.from_user.id}) завершил отправку данных по заказу из Китая.\n"
        f"Тариф: {tariff}",
    )


@dp.message(F.text == "Оплатить заказ")
async def request_payment(message: Message, bot: Bot) -> None:
    if await maintenance_guard(message):
        return
    if await banned_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    if not message.from_user:
        return

    user = get_user(message.from_user.id)
    if not user or (user.get("payment_status", "") or "") != "approved_waiting_payment":
        await message.answer("Сейчас оплата для этого заказа недоступна.")
        return

    update_user_fields(message.from_user.id, payment_status="requested_requisites")
    await message.answer("Запрос на оплату отправлен админу. Ожидай реквизиты.")
    await bot.send_message(
        ADMIN_ID,
        f"Пользователь {message.from_user.full_name} "
        f"(ID: {message.from_user.id}) запросил оплату.\n"
        "Отправь реквизиты командой:\n"
        f"/payreq {message.from_user.id} РЕКВИЗИТЫ"
    )


@dp.message(F.text == "Назад")
async def back_to_main_menu(message: Message) -> None:
    if await maintenance_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return
    update_user(message, tariff="", status="В главном меню")
    if message.from_user:
        admin_excel_sessions.pop(message.from_user.id, None)
        china_order_users.discard(message.from_user.id)
        admin_search_users.discard(message.from_user.id)
        paid_search_users.discard(message.from_user.id)
        tracking_lookup_users.discard(message.from_user.id)
    user_name = message.from_user.first_name if message.from_user else "друг"
    await send_main_menu(message, user_name)


@dp.message()
async def handle_messages(message: Message, bot: Bot) -> None:
    if await maintenance_guard(message):
        return
    if await banned_guard(message):
        return
    if is_rate_limited(message):
        await maybe_send_rate_limit_notice(message)
        return

    update_user(message)

    if not message.from_user:
        return

    session = admin_excel_sessions.get(message.from_user.id)
    if is_admin(message) and session and session.get("mode") == "excel_order":
        ensure_cargo_dirs()
        step = session.get("step")
        if step == "photo":
            if not message.photo:
                await message.answer("Отправь фото товара.")
                return
            photo = message.photo[-1]
            photo_path = CARGO_PHOTOS_DIR / f"{message.from_user.id}_{int(time.time())}.jpg"
            await bot.download(photo, destination=str(photo_path))
            session["data"]["photo_path"] = photo_path
            session["step"] = "name"
            await message.answer("Теперь отправь название товара.")
            return
        if step == "name":
            session["data"]["name"] = (message.text or "").strip()
            session["step"] = "size"
            await message.answer("Теперь отправь размер.")
            return
        if step == "size":
            session["data"]["size"] = (message.text or "").strip()
            session["step"] = "color"
            await message.answer("Теперь отправь цвет.")
            return
        if step == "color":
            session["data"]["color"] = (message.text or "").strip()
            session["step"] = "link"
            await message.answer("Теперь отправь ссылку на товар.")
            return
        if step == "link":
            session["data"]["link"] = (message.text or "").strip()
            session["step"] = "price"
            await message.answer("Теперь отправь цену за единицу.")
            return
        if step == "price":
            session["data"]["price"] = (message.text or "").strip()
            session["step"] = "quantity"
            await message.answer("Теперь отправь количество.")
            return
        if step == "quantity":
            session["data"]["quantity"] = (message.text or "").strip()
            session["step"] = "delivery"
            await message.answer("Теперь отправь доставку по Китаю.")
            return
        if step == "delivery":
            session["data"]["delivery"] = (message.text or "").strip()
            row = append_product_to_cargo(session["file_path"], session["data"])
            file_name = Path(session["file_path"]).name
            admin_excel_sessions[message.from_user.id] = {
                "mode": "excel_order",
                "step": "photo",
                "file_path": session["file_path"],
                "data": {},
            }
            await message.answer(
                f"Товар добавлен в {file_name}, строка {row}.\n"
                "Можешь сразу отправить фото следующего товара или использовать /excelview current."
            )
            return

    if message.from_user.id in admin_search_users and is_admin(message):
        query = (message.text or message.caption or "").strip()
        if not query:
            await message.answer("Для поиска пришли название вещи текстом или фото с подписью.")
            return

        await message.answer("Ищу по твоим ссылкам, подожди немного...")
        results = await search_item_in_sources(query)
        await message.answer(build_result_text(results))
        return

    if message.from_user.id in paid_search_users:
        user = get_user(message.from_user.id)
        credits = int((user or {}).get("search_access", 0) or 0)
        if credits <= 0:
            paid_search_users.discard(message.from_user.id)
            await message.answer(
                "У тебя закончились попытки поиска.\n"
                "Чтобы получить новые, напиши в техподдержку: @ichov",
                reply_markup=current_main_keyboard(message),
            )
            return

        if message.photo:
            await message.answer("Поиск по фото доступен только админу. Пришли название вещи текстом.")
            return

        query = (message.text or "").strip()
        if not query:
            await message.answer("Для поиска пришли название вещи текстом.")
            return

        remaining = consume_search_credit(message.from_user.id)
        if remaining is None:
            paid_search_users.discard(message.from_user.id)
            await message.answer(
                "У тебя закончились попытки поиска.\n"
                "Чтобы получить новые, напиши в техподдержку: @ichov",
                reply_markup=current_main_keyboard(message),
            )
            return

        update_user(message, status="Ищет вещь в Китае")
        await message.answer("Ищу по твоим ссылкам, подожди немного...")
        results = await search_item_in_sources(query)
        await message.answer(
            build_result_text(results)
            + f"\n\nОсталось попыток поиска: {remaining}"
        )
        if remaining <= 0:
            paid_search_users.discard(message.from_user.id)
        return

    if message.from_user.id in tracking_lookup_users:
        tracking_code = (message.text or "").strip().upper()
        _, user = get_user_by_tracking_code(tracking_code)
        if not tracking_code:
            await message.answer("Отправь трек-код текстом.")
            return
        if not user:
            await message.answer("Заказ с таким кодом не найден.")
            return

        await message.answer(
            f"Трек-код: {tracking_code}\n"
            f"Текущий этап: {user.get('tracking_stage', '') or 'Этап пока не назначен'}",
            reply_markup=info_keyboard,
        )
        return

    current_user = get_user(message.from_user.id)
    if current_user and (current_user.get("payment_status", "") or "") == "waiting_payment":
        await bot.forward_message(ADMIN_ID, message.chat.id, message.message_id)
        await bot.send_message(
            ADMIN_ID,
            f"Пользователь {message.from_user.full_name} "
            f"(ID: {message.from_user.id}) отправил подтверждение оплаты.\n"
            f"Подтверди оплату командой:\n/paid {message.from_user.id}"
        )
        await message.answer("Подтверждение оплаты отправлено админу. Ожидай проверку.")
        return

    if is_admin(message):
        if message.text:
            await message.answer(
                "Для работы используй команды:\n"
                "/send user_id текст\n"
                "/done user_id\n"
                "/cancel user_id\n"
                "/orders\n"
                "/active\n"
                "/tracks\n"
                "/orderno номер\n"
                "/today\n"
                "/recent\n"
                "/search запрос\n"
                "/grantsearch user_id [кол-во]\n"
                "/setsearch user_id кол-во\n"
                "/revokesearch user_id\n"
                "/searchlist\n"
                "/status user_id статус\n"
                "/owner код\n"
                "/broadcast текст\n"
                "/maintenance on|off\n"
                "/find название вещи\n"
                "/track код\n"
                "/trackset код этап\n"
                "/vykup код\n"
                "/sklad код\n"
                "/otpravlen код\n"
                "/rf код"
            )
        return

    if message.from_user.id in china_order_users:
        update_user(message, country="Китай", status="Отправляет данные")
        await notify_admin(bot, message)
        await bot.forward_message(ADMIN_ID, message.chat.id, message.message_id)
        await message.answer(
            "Принял. Если все отправил, нажми кнопку 'Я все скинул'.",
            reply_markup=china_submit_keyboard,
        )


async def main() -> None:
    ensure_cargo_dirs()
    init_db()
    migrate_legacy_users()
    purge_canceled_orders()
    bot = Bot(token=TOKEN)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

